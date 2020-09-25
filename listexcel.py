from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
#openpyxl


root = Tk()
root.title('Codemy.com - Listbox Excel')
root.iconbitmap('c:/gui/codemy.ico')
root.geometry("400x300")

#my_list = ["One", "Two", "Three", "Four", "Five"]

def select(e):
	my_label.config(text=my_box.get(ANCHOR))

my_box = Listbox(root, width=45)
my_box.pack(pady=20)

#for item in my_list:
#	my_box.insert(END, item)
wb = load_workbook('name_color.csv')
ws = wb.active

col_a = ws["A"]
col_b = ws["B"]

for item in col_a:
	my_box.insert(END, item.value)

my_label = Label(root, text='Select Item...', font=("Helvetica", 18))
my_label.pack(pady=20)

my_box.bind("<ButtonRelease-1>", select)

root.mainloop()