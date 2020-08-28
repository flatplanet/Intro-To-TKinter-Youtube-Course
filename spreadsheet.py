from tkinter import *
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root = Tk()
root.title('Codemy.com - OpenPyXL')
root.iconbitmap('c:/gui/codemy.ico')
root.geometry("500x800")

# Create workbook instance
wb = Workbook()
# Load existing workbook
wb = load_workbook('pizza.xlsx')
# Create active worksheet
ws = wb.active

# Create variable for Column A
column_a = ws['A']
column_b = ws['B']

def get_a():
	list = ''
	for cell in column_a:
		list = f'{list + str(cell.value)}\n'
		
	label_a.config(text=list)

def get_b():
	list = ''
	for cell in column_b:
		list = f'{list + str(cell.value)}\n'
		
	label_b.config(text=list)

ba = Button(root, text="Get Column A", command=get_a)
ba.pack(pady=20)

label_a = Label(root, text="")
label_a.pack(pady=20)


bb = Button(root, text="Get Column B", command=get_b)
bb.pack(pady=20)

label_b = Label(root, text="")
label_b.pack(pady=20)

# Add fred to A8 and B8
ws['A8'] = "Fred"
ws['B8'] = "Cheese"


# Save new file
wb.save('pizza2.xlsx')



root.mainloop()